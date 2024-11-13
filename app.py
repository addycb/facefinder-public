from flask import Flask, request, jsonify, render_template, send_file
from werkzeug.utils import secure_filename
import os
import main
import time
import random
import requests
from io import BytesIO
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
import base64
from PIL import Image
import sys



app = Flask(__name__)
def save_json_to_file(data):
    """Save JSON data to 'savedson.txt'."""
    filename = 'savedson.txt'
    try:
        with open(filename, 'w') as file:
            json.dump(data, file, indent=4)
        print(f"Data successfully saved to {filename}")
    except Exception as e:
        print(f"An error occurred while saving the data: {e}")



def save_to_excel(nodeDataMap,edges):
    #print leaf node statsu of each node
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Search Report"
    # Set the header
    headers = ['Label', 'Representative Image','Node Id','Parent Node ID', 'Level', 'Images', 'EXIF Data', 'Notes']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        worksheet[f'{col_letter}1'] = header
    worksheet.column_dimensions['B'].width = 114
    # Populate the rows with data
    row_num = 2
    #print("parentmap making...")
    edges=json.loads(edges)
    parent_map = {edge['to']: edge['from'] for edge in edges}
    nodeDataMap=json.loads(nodeDataMap)
    #print("worksheet start!")
    for node_id, node_data in nodeDataMap.items():
        if(node_data.get('searchable', True)):
           pass
        else:
            worksheet[f'A{row_num}'] = node_data.get('label', 'No Label')
            #print(node_data.get('label', 'No Label'))
            #Add representative image to cell from base64 string
            #print("adding img to excel")
            reprimage= node_data.get('imagedata', 'No Image')
            #print("imagedata: ")
            #print(reprimage)
            reprimage=str(reprimage)
            reprimage=reprimage.split(",")[1]
            reprimage=base64.b64decode(reprimage)
            #print("base64 str len:",len(reprimage))
            #with open("imageToSave", "w") as fh:
            #    fh.write(reprimage)
            #print(f"Base64 string saved")
            #print("decoded img")
            image=Image.open(BytesIO(reprimage))
            #print("image opened in bytesio")
            image_format = image.format if image.format else 'PNG'
            #print(image_format)
            img_stream=BytesIO()
            max_size = (150, 150)
            image.thumbnail(max_size)
            image.save(img_stream, format=image_format)
            img_stream.seek(0)
            #print("image saved")
            reprimage=ExcelImage(img_stream)
            worksheet.add_image(reprimage, f'B{row_num}')
            #print("image added")
            worksheet.row_dimensions[row_num].height = 114
            worksheet[f'C{row_num}'] = node_id
            worksheet[f'D{row_num}'] = parent_map.get(int(node_id), 'No Parent Node')
            worksheet[f'E{row_num}'] = node_data['level']
            # Handle images and EXIF data
            images = []
            exif_data = []
            #if node is not a leaf node
            if(node_data.get('leaf', True)):
                #print("leaf node")  
                pass
            else:
                #print("not leaf node")
                for image in node_data['images']:
                    images.append(image['imagelink'])
                    exif_data.append(image['exifdata'])
            worksheet[f'F{row_num}'] = ', '.join(images)
            worksheet[f'G{row_num}'] = ', '.join(exif_data)
            # Handle notes
            worksheet[f'H{row_num}'] = node_data['notes']
            row_num += 1
            
    # Save to a BytesIO object
    excel_output = BytesIO()
    workbook.save(excel_output)
    excel_output.seek(0)
    
    return excel_output

@app.route('/export', methods=['POST'])
def export_excel():
    try:
        uploaded=request.json
        nodeDataMap=uploaded.get('nodeDataMap')
        edges=uploaded.get('edges')
        #print(nodeDataMap)
        #save_json_to_file(nodeDataMap)
        excel_output = save_to_excel(nodeDataMap,edges)

        return send_file(excel_output, as_attachment=True, download_name='document.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        print(str(e))
        return jsonify({'error': str(e)}), 500

# Set up an upload folder
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Create the upload folder if it doesn't exist
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

@app.route('/')
def index():
    main.myproxies.log_visitor(request.remote_addr)
    return render_template('newindex.html')


@app.route('/upload', methods=['POST'])
def upload_image():
    if 'image' not in request.files:
        return 
    file = request.files['image']
    if file:
        # Secure the filename and save it to the upload folder
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        #log the image
        main.myproxies.log_image(request.remote_addr,filename)
        # Process the image or perform any required logic here
        # For example, you might want to analyze the image or extract metadata
        
        #Make sure json returned is in format
        """
        images: an array of images for the node
            - imagelink: the link to the image
            - otherpeople: an array of links to other people in the image
            - contextlinks: the links to the image context
            - exifdata: the EXIF data for the image
        """
        jsontoreturn=main.search(filepath)
        #Get rid of the file
        #os.remove(filepath)
        return jsontoreturn


if __name__ == '__main__':
    app.run(debug=True)